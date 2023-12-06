from flask import Flask, render_template, request, redirect, session, make_response, url_for
import psutil
import cpuinfo
import socket
import platform
import mysql.connector
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

app = Flask(__name__)
app.secret_key = 'Chester'

# Datos de conexión a la base de datos MySQL
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="",
    database="sistemapsutil"
)


@app.route('/')
def home():
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
    # Obtener los datos del formulario de inicio de sesión
    username = request.form['username']
    password = request.form['password']

    # Validar el inicio de sesión
    cursor = mydb.cursor()
    cursor.execute(
        "SELECT * FROM users WHERE username=%s AND password=%s", (username, password))
    user = cursor.fetchone()

    if user is not None:
        return redirect('/index')
    else:
        return render_template('login.html', error=True)


@app.route('/index', methods=['GET'])
def index():
    return render_template('index.html')


@app.route('/escanear', methods=['GET', 'POST'])
def escanear():
    if request.method == 'POST':
        # Obtener la información del sistema con psutil
        nombre_pc = platform.node()
        sistema_operativo = platform.system()
        version_sistema_operativo = platform.release()
        procesador = cpuinfo.get_cpu_info()['brand_raw']
        arquitectura_procesador = platform.machine()
        procesadores_fisicos = psutil.cpu_count(logical=False)
        total_procesadores = psutil.cpu_count(logical=True)
        frecuencia_base_procesador = psutil.cpu_freq().current
        uso_total_cpu = psutil.cpu_percent(interval=1, percpu=False)
        memoria_total = round(psutil.virtual_memory().total /
                              (1024*1024*1024), 2)  # En GB
        memoria_disponible = round(
            psutil.virtual_memory().available/(1024*1024*1024), 2)  # En GB
        memoria_usada = round(psutil.virtual_memory().used /
                              (1024*1024*1024), 2)  # En GB
        porcentaje_uso_memoria = psutil.virtual_memory().percent
        capacidad_disco_total = psutil.disk_usage('/').total / (1024**3)
        particion_disco = psutil.disk_partitions()[0].device
        disco_usado = psutil.disk_usage('/').used / (1024**3)
        disco_disponible = psutil.disk_usage('/').free / (1024**3)
        porcentaje_uso_disco = psutil.disk_usage('/').percent
        direccion_mac = psutil.net_if_addrs()['Ethernet'][0].address
        direccion_ip = socket.gethostbyname(socket.gethostname())

        # Guardar la información en la base de datos
        mycursor = mydb.cursor()
        sql = "INSERT INTO datos_pc (nombre_pc, sistema_operativo, version_sistema_operativo, procesador, arquitectura_procesador, procesadores_fisicos, total_procesadores, frecuencia_base_procesador, uso_total_cpu, memoria_total, memoria_disponible, memoria_usada, porcentaje_uso_memoria, capacidad_disco_total, particion_disco, disco_usado, disco_disponible, porcentaje_uso_disco, direccion_mac, direccion_ip) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        val = (nombre_pc, sistema_operativo, version_sistema_operativo, procesador, arquitectura_procesador, procesadores_fisicos, total_procesadores, frecuencia_base_procesador, uso_total_cpu,
               memoria_total, memoria_disponible, memoria_usada, porcentaje_uso_memoria, capacidad_disco_total, particion_disco, disco_usado, disco_disponible, porcentaje_uso_disco, direccion_mac, direccion_ip)
        mycursor.execute(sql, val)
        mydb.commit()

    # Renderizar la plantilla con los datos escaneados
    return render_template('escanear.html', nombre_pc=nombre_pc, sistema_operativo=sistema_operativo,
                           version_sistema_operativo=version_sistema_operativo, procesador=procesador,
                           arquitectura_procesador=arquitectura_procesador, procesadores_fisicos=procesadores_fisicos,
                           total_procesadores=total_procesadores, frecuencia_base_procesador=frecuencia_base_procesador,
                           uso_total_cpu=uso_total_cpu, memoria_total=memoria_total, memoria_disponible=memoria_disponible,
                           memoria_usada=memoria_usada, porcentaje_uso_memoria=porcentaje_uso_memoria, capacidad_disco_total=capacidad_disco_total, particion_disco=particion_disco,
                           disco_usado=disco_usado, disco_disponible=disco_disponible,
                           porcentaje_uso_disco=porcentaje_uso_disco, direccion_mac=direccion_mac, direccion_ip=direccion_ip)


@app.route('/volver', methods=['POST'])
def volver():
    # Redirigir a la página de inicio (index.html)
    return redirect('/index')


@app.route('/logout', methods=['GET'])
def logout():
    # Elimina el nombre de usuario de la sesión
    session.pop('username', None)
    # Redirige a la página de inicio de sesión después de cerrar sesión
    return redirect('/')


@app.route('/inventario', methods=['POST'])
def inventario():
    # Obtener los datos del inventario de la base de datos
    mycursor = mydb.cursor()
    sql = "SELECT * FROM datos_pc"
    mycursor.execute(sql)
    data = mycursor.fetchall()

    # Renderizar la plantilla con los datos del inventario
    return render_template('inventario.html', data=data)


@app.route('/eliminar_pc/<int:pc_id>')
def eliminar_pc(pc_id):
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='sistemapsutil'
    )
    cursor = conn.cursor()
    # Eliminar la fila de la variable data
    cursor.execute('SELECT * FROM datos_pc')
    data = cursor.fetchall()
    for i in range(len(data)):
        if data[i][0] == pc_id:
            data.pop(i)
            break

    # Renderizar la plantilla actualizada con los datos
    return render_template('inventario.html', data=data)


@app.route('/exportar-xlsx')
def exportar_xlsx():
    # Obtener los datos del historial de la base de datos
    mycursor = mydb.cursor()
    sql = "SELECT * FROM datos_pc"
    mycursor.execute(sql)
    data = mycursor.fetchall()

    # Crear el archivo xlsx
    wb = Workbook()
    ws = wb.active

    # Agregar el título a la hoja de cálculo
    ws.title = "Inventario de Escaneos"
    ws['A1'] = "Id"
    ws['B1'] = "Nombre de PC"
    ws['C1'] = "Sistema Operativo"
    ws['D1'] = "Versión del S.O"
    ws['E1'] = "Procesador"
    ws['F1'] = "Total de Nucleos"
    ws['G1'] = "Memoria Total"
    ws['H1'] = "Disco Total"
    ws['I1'] = "Ubicacion"
    bold_font = Font(bold=True)
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')

    # Agregar los datos a la hoja de cálculo
    for i, row in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        ws.cell(row=i, column=3, value=row[2])
        ws.cell(row=i, column=4, value=row[3])
        ws.cell(row=i, column=5, value=row[4])
        ws.cell(row=i, column=6, value=row[7])
        ws.cell(row=i, column=7, value=row[10])
        ws.cell(row=i, column=8, value=row[14])

    # Guardar el archivo xlsx
    filename = "Inventario.xlsx"
    wb.save(filename)

    # Devolver el archivo xlsx como respuesta
    with open(filename, 'rb') as f:
        response = make_response(f.read())
    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-Disposition'] = 'attachment; filename=' + filename
    return response


productos = [
    {"id": 1, "nombre": "GPU NVIDIA GeForce RTX 3080", "precio": "3,031,820"},
    {"id": 2, "nombre": "Procesador Intel Core i9-11900K", "precio": "2,091,810"},
    {"id": 3, "nombre": "Memoria RAM Corsair Vengeance RGB Pro 32GB", "precio": "569,142"},
    {"id": 4, "nombre": "Disco Duro SSD Samsung 1TB", "precio": "493,822"},
    {"id": 5, "nombre": "Monitor Dell UltraSharp 27 Pulgadas", "precio": "1,139,622"},
    {"id": 6, "nombre": "Teclado mecánico Razer BlackWidow Elite", "precio": "646,142"},
    {"id": 7, "nombre": "Mouse Logitech G Pro Wireless", "precio": "569,142"},
    {"id": 8, "nombre": "Auriculares Sony WH-1000XM4", "precio": "1,139,622"},
    {"id": 9, "nombre": "Impresora HP LaserJet Pro", "precio": "759,748"},
    {"id": 10, "nombre": "Webcam Logitech C920", "precio": "303,182"},
    {"id": 11, "nombre": "Silla de oficina ergonómica", "precio": "949,748"},
    {"id": 12, "nombre": "Laptop Dell XPS 13", "precio": "4,551,822"},
    {"id": 13, "nombre": "Router inalámbrico ASUS RT-AX88U", "precio": "1,139,622"},
    {"id": 14, "nombre": "Altavoces Logitech Z623", "precio": "455,182"},
    {"id": 15, "nombre": "Tablet Apple iPad Pro 12.9 Pulgadas", "precio": "3,797,476"}
]

@app.route('/compra_productos', methods=['GET', 'POST'])
def compra_productos():
    if request.method == 'POST':
        producto_id = int(request.form['producto_id'])
        accion = request.form['accion']

        # Lógica para cotizar
        if accion == 'cotizar':
            # Aquí puedes agregar lógica específica para la cotización
            producto = obtener_producto_por_id(producto_id)
            if producto:
                return redirect(url_for('cotizar', producto_id=producto_id))
            else:
                # Manejo de error si no se encuentra el producto
                return render_template('error.html', mensaje='Producto no encontrado')

        # Lógica para comprar
        elif accion == 'comprar':
            # Aquí puedes agregar lógica específica para la compra
            producto = obtener_producto_por_id(producto_id)
            if producto:
                return redirect(url_for('comprar', producto_id=producto_id))
            else:
                # Manejo de error si no se encuentra el producto
                return render_template('error.html', mensaje='Producto no encontrado')

    return render_template('compra_productos.html', productos=productos)

@app.route('/cotizar/<int:producto_id>')
def cotizar(producto_id):
    # Aquí deberías obtener el producto correspondiente según el producto_id
    producto = obtener_producto_por_id(producto_id)

    if producto:
        return render_template('cotizar.html', producto=producto)
    else:
        # Manejo de error si no se encuentra el producto
        return render_template('error.html', mensaje='Producto no encontrado')

@app.route('/comprar/<int:producto_id>')
def comprar(producto_id):
    # Aquí también deberías obtener el producto correspondiente según el producto_id
    producto = obtener_producto_por_id(producto_id)

    if producto:
        return render_template('comprar.html', producto=producto)
    else:
        # Manejo de error si no se encuentra el producto
        return render_template('error.html', mensaje='Producto no encontrado')

# Función de ejemplo para obtener un producto por su ID
def obtener_producto_por_id(producto_id):
    # Lógica para buscar el producto en la lista de productos
    for producto in productos:
        if producto['id'] == producto_id:
            return producto
    return None

from flask import render_template

@app.route('/reporte_ventas')
def reporte_ventas():
    # Lógica para obtener datos de ventas desde la base de datos (puedes ajustar según tus necesidades)
    # Aquí deberías tener una consulta a tu base de datos para obtener los datos de ventas

    # Ejemplo de datos de ventas (puedes ajustarlo según tus necesidades)
    datos_ventas = [
        {"id": 1, "producto": "Laptop", "cantidad": 5, "total": 20000000},
        {"id": 2, "producto": "Teclado", "cantidad": 10, "total": 4000000},
        {"id": 3, "producto": "Monitor", "cantidad": 8, "total": 8000000},
        {"id": 4, "producto": "Mouse", "cantidad": 15, "total": 3000000},
        {"id": 5, "producto": "Impresora", "cantidad": 3, "total": 12000000},
        {"id": 6, "producto": "Tablet", "cantidad": 6, "total": 7200000},
        {"id": 7, "producto": "Altavoces", "cantidad": 12, "total": 4800000},
        {"id": 8, "producto": "Disco Duro", "cantidad": 7, "total": 14000000},
        {"id": 9, "producto": "Memoria RAM", "cantidad": 20, "total": 1600000},
        {"id": 10, "producto": "Tarjeta Gráfica", "cantidad": 4, "total": 12000000},
        {"id": 11, "producto": "Router", "cantidad": 9, "total": 3600000},
        {"id": 12, "producto": "Webcam", "cantidad": 2, "total": 3200000},
        {"id": 13, "producto": "Auriculares", "cantidad": 11, "total": 4400000},
        {"id": 14, "producto": "Cargador", "cantidad": 15, "total": 3000000},
        {"id": 15, "producto": "Batería", "cantidad": 3, "total": 4800000},
        {"id": 16, "producto": "Mochila", "cantidad": 6, "total": 2400000},
        {"id": 17, "producto": "Estuche para Laptop", "cantidad": 8, "total": 3200000},
        {"id": 18, "producto": "Micrófono", "cantidad": 4, "total": 1600000},
        {"id": 19, "producto": "Cámara de Seguridad", "cantidad": 7, "total": 8400000},
        {"id": 20, "producto": "Silla Gamer", "cantidad": 2, "total": 1600000},
        {"id": 21, "producto": "Soporte para Monitor", "cantidad": 5, "total": 2000000},
        {"id": 22, "producto": "Lámpara LED", "cantidad": 3, "total": 1200000},
        {"id": 23, "producto": "Cable HDMI", "cantidad": 10, "total": 5000000},
        {"id": 24, "producto": "Hub USB", "cantidad": 6, "total": 2400000},
        {"id": 25, "producto": "Tarjeta de Sonido", "cantidad": 1, "total": 6000000},
    ]

    # Calcular el total neto y el total final de las ventas
    total_neto = sum(venta["total"] for venta in datos_ventas)
    impuestos = total_neto * 0.16  # Supongamos un impuesto del 16%
    total_final = total_neto + impuestos

    return render_template('reporte_ventas.html', datos_ventas=datos_ventas, total_neto=total_neto, impuestos=impuestos, total_final=total_final)

# Datos de ventas
datos_ventas = [
    {"id": 1, "producto": "Laptop", "cantidad": 5, "total": 20000000},
    {"id": 2, "producto": "Teclado", "cantidad": 10, "total": 4000000},
    {"id": 3, "producto": "Monitor", "cantidad": 8, "total": 8000000},
    {"id": 4, "producto": "Mouse", "cantidad": 15, "total": 3000000},
    {"id": 5, "producto": "Impresora", "cantidad": 3, "total": 12000000},
    {"id": 6, "producto": "Tablet", "cantidad": 6, "total": 7200000},
    {"id": 7, "producto": "Altavoces", "cantidad": 12, "total": 4800000},
    {"id": 8, "producto": "Disco Duro", "cantidad": 7, "total": 14000000},
    {"id": 9, "producto": "Memoria RAM", "cantidad": 20, "total": 1600000},
    {"id": 10, "producto": "Tarjeta Gráfica", "cantidad": 4, "total": 12000000},
    # ... más datos de ventas
]

# Datos de productos
datos_productos = [
    {"id": 1, "nombre": "Laptop", "stock": 20, "precio_unitario": 80000000},
    {"id": 2, "nombre": "Teclado", "stock": 50, "precio_unitario": 20000000},
    {"id": 3, "nombre": "Monitor", "stock": 30, "precio_unitario": 40000000},
    {"id": 4, "nombre": "Mouse", "stock": 25, "precio_unitario": 6000000},
    {"id": 5, "nombre": "Impresora", "stock": 10, "precio_unitario": 12000000},
    {"id": 6, "nombre": "Tablet", "stock": 15, "precio_unitario": 12000000},
    {"id": 7, "nombre": "Altavoces", "stock": 20, "precio_unitario": 4000000},
    {"id": 8, "nombre": "Disco Duro", "stock": 12, "precio_unitario": 24000000},
    {"id": 9, "nombre": "Memoria RAM", "stock": 40, "precio_unitario": 2000000},
    {"id": 10, "nombre": "Tarjeta Gráfica", "stock": 8, "precio_unitario": 30000000},
    # ... más datos de productos
]

@app.route('/generar_informes')
def informes():
    # Lógica para obtener estadísticas
    total_ventas = sum(venta["total"] for venta in datos_ventas)
    productos_agotados = [producto["nombre"] for producto in datos_productos if producto["stock"] == 0]

    # Ventas
    total_ventas = sum(venta["total"] for venta in datos_ventas)
    ingreso_promedio_por_venta = total_ventas / len(datos_ventas) if datos_ventas else 0
    productos_mas_vendidos = [venta["producto"] for venta in sorted(datos_ventas, key=lambda x: x["cantidad"], reverse=True)[:5]]

    # Productos
    inventario_actual = sum(producto["stock"] for producto in datos_productos)
    productos_mas_populares = [producto["nombre"] for producto in sorted(datos_productos, key=lambda x: x["stock"], reverse=True)[:5]]

    # Finanzas
    costos_operativos = 5000  # Reemplaza con tus datos reales
    margen_beneficio = total_ventas - costos_operativos
    comparacion_ingresos_gastos = {"Ingresos": total_ventas, "Gastos": costos_operativos}
    flujo_efectivo = total_ventas - costos_operativos

    # Más métricas...

    return render_template('informes.html', 
                           total_ventas=total_ventas, 
                           ingreso_promedio_por_venta=ingreso_promedio_por_venta, 
                           productos_mas_vendidos=productos_mas_vendidos,
                           inventario_actual=inventario_actual,
                           productos_mas_populares=productos_mas_populares,
                           margen_beneficio=margen_beneficio,
                           comparacion_ingresos_gastos=comparacion_ingresos_gastos,
                           flujo_efectivo=flujo_efectivo,
                           datos_ventas=datos_ventas,
                           datos_productos=datos_productos)

@app.route('/clientes_y_proveedores')
def clientes_y_proveedores():
    # Datos de clientes
    clientes = [
        {"nombre": "Juan Pérez", "direccion": "Calle 123, Ciudad 1", "telefono": "123-456-7890", "correo": "juan@example.com", "historial_compras": ["Laptop", "Teclado", "Monitor"]},
        {"nombre": "María Gómez", "direccion": "Calle 456, Ciudad 2", "telefono": "987-654-3210", "correo": "maria@example.com", "historial_compras": ["Impresora", "Tablet", "Altavoces"]},
        {"nombre": "Carlos Rodríguez", "direccion": "Calle 789, Ciudad 3", "telefono": "111-222-3333", "correo": "carlos@example.com", "historial_compras": ["Memoria RAM", "Tarjeta Gráfica", "Router"]},
        {"nombre": "Laura Martínez", "direccion": "Calle 012, Ciudad 4", "telefono": "444-555-6666", "correo": "laura@example.com", "historial_compras": ["Cámara de Seguridad", "Silla Gamer", "Lámpara LED"]},
        {"nombre": "Roberto Sánchez", "direccion": "Calle 345, Ciudad 5", "telefono": "777-888-9999", "correo": "roberto@example.com", "historial_compras": ["Cable HDMI", "Hub USB", "Tarjeta de Sonido"]}
    ]

    # Datos de proveedores
    proveedores = [
        {"nombre": "TechPro", "contacto": "Ana López", "terminos_pago": "30 días", "historial_compras": ["Procesadores", "Discos Duros", "Fuentes de Poder"]},
        {"nombre": "ElectroMega", "contacto": "Javier Ramírez", "terminos_pago": "45 días", "historial_compras": ["Pantallas LCD", "Cámaras Web", "Auriculares"]},
        {"nombre": "GadgetLand", "contacto": "Claudia Herrera", "terminos_pago": "60 días", "historial_compras": ["Baterías", "Cargadores", "Soportes para Monitor"]},
        {"nombre": "AccessoriesWorld", "contacto": "Luis Torres", "terminos_pago": "30 días", "historial_compras": ["Estuches para Laptop", "Micrófonos", "Mochilas"]},
        {"nombre": "ConnectivityHub", "contacto": "Marta Díaz", "terminos_pago": "45 días", "historial_compras": ["Tarjetas de Red", "Adaptadores USB", "Lectores de Tarjetas"]}
    ]

    return render_template('clientes_y_proveedores.html', clientes=clientes, proveedores=proveedores)

if __name__ == '__main__':
    app.run(debug=True)
if __name__ == '__main__':
    app.run(debug=True)